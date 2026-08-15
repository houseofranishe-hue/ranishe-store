"""
House of Ranishè — store backend (v2)
=====================================
Full shop system: storefront API + admin CRUD + reports (incl. P&L).

Database:
  * If DATABASE_URL is set (Render Postgres), uses Postgres  -> data is PERMANENT.
  * Otherwise falls back to local SQLite (for testing on your computer).

Env vars to set in Render:
  DATABASE_URL   -> auto-provided when you attach a Render Postgres
  ADMIN_TOKEN    -> your secret admin password
  UPI_ID         -> your UPI id shown to customers
  WHATSAPP       -> your WhatsApp number, e.g. 919167629547
"""

import os
import json
import uuid
from datetime import datetime, timezone
from typing import List, Optional

from fastapi import FastAPI, HTTPException, Header, Request
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field

ADMIN_TOKEN = os.environ.get("ADMIN_TOKEN", "change-me")
DATABASE_URL = os.environ.get("DATABASE_URL", "")
PUBLIC_CONFIG = {
    "upi_id": os.environ.get("UPI_ID", "anishmathew131@okicici"),
    "whatsapp": os.environ.get("WHATSAPP", "919167629547"),
    "store_name": "House of Ranishe",
}

USING_PG = DATABASE_URL.startswith("postgres")

if USING_PG:
    import psycopg2
    import psycopg2.extras
    def get_conn():
        return psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
else:
    import sqlite3
    DB_PATH = os.environ.get("DB_PATH", "store.db")
    def get_conn():
        c = sqlite3.connect(DB_PATH)
        c.row_factory = sqlite3.Row
        return c

app = FastAPI(title="House of Ranishè Store")


def q(sql: str) -> str:
    return sql if USING_PG else sql.replace("%s", "?")


def init_db():
    conn = get_conn(); cur = conn.cursor()
    if USING_PG:
        cur.execute("""CREATE TABLE IF NOT EXISTS categories (name TEXT PRIMARY KEY, sort_order INTEGER DEFAULT 0, image_url TEXT DEFAULT '')""")
        cur.execute("""CREATE TABLE IF NOT EXISTS products (
            sku TEXT PRIMARY KEY, name TEXT NOT NULL, category TEXT NOT NULL,
            cost_price INTEGER NOT NULL DEFAULT 0, price INTEGER NOT NULL, sale_price INTEGER NOT NULL,
            stock INTEGER NOT NULL DEFAULT 0, image_url TEXT DEFAULT '', description TEXT DEFAULT '',
            archived BOOLEAN DEFAULT FALSE)""")
        cur.execute("""CREATE TABLE IF NOT EXISTS orders (
            id TEXT PRIMARY KEY, created_at TEXT NOT NULL, customer_name TEXT NOT NULL,
            phone TEXT NOT NULL, address TEXT NOT NULL, items_json TEXT NOT NULL,
            total INTEGER NOT NULL, cost_total INTEGER NOT NULL DEFAULT 0, status TEXT NOT NULL DEFAULT 'new')""")
        cur.execute("""CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT DEFAULT '')""")
        cur.execute("""CREATE TABLE IF NOT EXISTS analytics_events (
            id SERIAL PRIMARY KEY, created_at TEXT NOT NULL, kind TEXT NOT NULL, sku TEXT)""")
        cur.execute("ALTER TABLE categories ADD COLUMN IF NOT EXISTS image_url TEXT DEFAULT ''")
    else:
        cur.execute("""CREATE TABLE IF NOT EXISTS categories (name TEXT PRIMARY KEY, sort_order INTEGER DEFAULT 0, image_url TEXT DEFAULT '')""")
        cur.execute("""CREATE TABLE IF NOT EXISTS products (
            sku TEXT PRIMARY KEY, name TEXT NOT NULL, category TEXT NOT NULL,
            cost_price INTEGER NOT NULL DEFAULT 0, price INTEGER NOT NULL, sale_price INTEGER NOT NULL,
            stock INTEGER NOT NULL DEFAULT 0, image_url TEXT DEFAULT '', description TEXT DEFAULT '',
            archived INTEGER DEFAULT 0)""")
        cur.execute("""CREATE TABLE IF NOT EXISTS orders (
            id TEXT PRIMARY KEY, created_at TEXT NOT NULL, customer_name TEXT NOT NULL,
            phone TEXT NOT NULL, address TEXT NOT NULL, items_json TEXT NOT NULL,
            total INTEGER NOT NULL, cost_total INTEGER NOT NULL DEFAULT 0, status TEXT NOT NULL DEFAULT 'new')""")
        cur.execute("""CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT DEFAULT '')""")
        cur.execute("""CREATE TABLE IF NOT EXISTS analytics_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT NOT NULL, kind TEXT NOT NULL, sku TEXT)""")
        # migrate older sqlite DBs created before image_url existed on categories
        cur.execute("PRAGMA table_info(categories)")
        cols = [row[1] for row in cur.fetchall()]
        if "image_url" not in cols:
            cur.execute("ALTER TABLE categories ADD COLUMN image_url TEXT DEFAULT ''")

    cur.execute("SELECT COUNT(*) AS c FROM products")
    row = cur.fetchone()
    count = row["c"] if isinstance(row, dict) else row[0]
    if count == 0 and os.path.exists("seed_products.json"):
        with open("seed_products.json", encoding="utf-8") as f:
            seed = json.load(f)
        for name, order in {"Earrings":0,"Chains":1,"Bracelets":2,"Gift Sets":3,"Under 499":4}.items():
            if USING_PG:
                cur.execute("INSERT INTO categories (name, sort_order) VALUES (%s,%s) ON CONFLICT DO NOTHING", (name, order))
            else:
                cur.execute("INSERT OR IGNORE INTO categories (name, sort_order) VALUES (?,?)", (name, order))
        for p in seed:
            cur.execute(q("""INSERT INTO products (sku,name,category,cost_price,price,sale_price,stock,image_url,description)
                             VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)"""),
                        (p["sku"], p["name"], p["category"], p.get("cost_price",0),
                         p["price"], p["sale_price"], p["stock"], p.get("image_url",""), p.get("description","")))
    conn.commit(); cur.close(); conn.close()


init_db()


def require_admin(token: Optional[str]):
    if token != ADMIN_TOKEN:
        raise HTTPException(401, "Invalid admin token.")


# ------------------------------------------------------- models
class ProductIn(BaseModel):
    sku: str = Field(min_length=1, max_length=40)
    name: str = Field(min_length=1, max_length=120)
    category: str = Field(min_length=1, max_length=60)
    cost_price: int = Field(ge=0)
    price: int = Field(ge=0)
    sale_price: int = Field(ge=0)
    stock: int = Field(ge=0)
    image_url: str = ""
    description: str = ""


class CategoryIn(BaseModel):
    name: str = Field(min_length=1, max_length=60)
    sort_order: int = 0
    image_url: str = ""


class SettingsIn(BaseModel):
    banner_text: str = ""
    banner_enabled: bool = False


class OrderItem(BaseModel):
    sku: str
    qty: int = Field(gt=0, le=50)


class OrderIn(BaseModel):
    customer_name: str = Field(min_length=2, max_length=100)
    phone: str = Field(min_length=8, max_length=15)
    address: str = Field(min_length=10, max_length=500)
    items: List[OrderItem]


# ------------------------------------------------------- storefront API
def get_setting(key: str, default: str = "") -> str:
    conn = get_conn(); cur = conn.cursor()
    cur.execute(q("SELECT value FROM settings WHERE key=%s"), (key,))
    row = cur.fetchone(); cur.close(); conn.close()
    if row is None:
        return default
    row = dict(row)
    return row["value"] if row["value"] is not None else default


@app.get("/api/config")
def config():
    out = dict(PUBLIC_CONFIG)
    out["banner_text"] = get_setting("banner_text", "")
    out["banner_enabled"] = get_setting("banner_enabled", "false") == "true"
    return out


@app.get("/api/categories")
def list_categories():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT name, sort_order, image_url FROM categories ORDER BY sort_order, name")
    rows = cur.fetchall(); cur.close(); conn.close()
    return [dict(r) for r in rows]


@app.get("/api/products")
def list_products():
    conn = get_conn(); cur = conn.cursor()
    af = "FALSE" if USING_PG else "0"
    cur.execute(f"SELECT * FROM products WHERE archived={af} ORDER BY category, name")
    rows = cur.fetchall(); cur.close(); conn.close()
    out = []
    for r in rows:
        d = dict(r); d.pop("cost_price", None); d.pop("archived", None); out.append(d)
    return out


@app.post("/api/orders")
def create_order(order: OrderIn):
    if not order.items:
        raise HTTPException(400, "Cart is empty.")
    conn = get_conn(); cur = conn.cursor()
    try:
        total = 0; cost_total = 0; detailed = []
        for item in order.items:
            cur.execute(q("SELECT * FROM products WHERE sku=%s"), (item.sku,))
            row = cur.fetchone()
            if row is None:
                raise HTTPException(400, f"Unknown product: {item.sku}")
            row = dict(row)
            if row["stock"] < item.qty:
                raise HTTPException(409, f"Only {row['stock']} left of {row['name']} — please adjust your cart.")
            total += row["sale_price"] * item.qty
            cost_total += row["cost_price"] * item.qty
            detailed.append({"sku":row["sku"],"name":row["name"],"qty":item.qty,"unit_price":row["sale_price"]})
        # delivery charge: Rs 60 for orders below Rs 999, free otherwise
        subtotal = total
        shipping = 60 if (0 < subtotal < 999) else 0
        total = subtotal + shipping
        for item in order.items:
            cur.execute(q("UPDATE products SET stock = stock - %s WHERE sku=%s"), (item.qty, item.sku))
        order_id = "RAN-" + uuid.uuid4().hex[:8].upper()
        cur.execute(q("""INSERT INTO orders (id,created_at,customer_name,phone,address,items_json,total,cost_total)
                         VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"""),
                    (order_id, datetime.now(timezone.utc).isoformat(timespec="seconds"),
                     order.customer_name.strip(), order.phone.strip(), order.address.strip(),
                     json.dumps(detailed), total, cost_total))
        conn.commit()
    finally:
        cur.close(); conn.close()
    return {"order_id": order_id, "total": total, "subtotal": subtotal, "shipping": shipping, "items": detailed}


# ------------------------------------------------------- admin: products
@app.get("/api/admin/products")
def admin_products(x_admin_token: Optional[str] = Header(None)):
    require_admin(x_admin_token)
    conn = get_conn(); cur = conn.cursor()
    af = "FALSE" if USING_PG else "0"
    cur.execute(f"SELECT * FROM products WHERE archived={af} ORDER BY category, name")
    rows = cur.fetchall(); cur.close(); conn.close()
    return [dict(r) for r in rows]


@app.post("/api/admin/products")
def upsert_product(p: ProductIn, x_admin_token: Optional[str] = Header(None)):
    require_admin(x_admin_token)
    conn = get_conn(); cur = conn.cursor()
    cur.execute(q("SELECT sku FROM products WHERE sku=%s"), (p.sku,))
    exists = cur.fetchone()
    if exists:
        cur.execute(q("""UPDATE products SET name=%s,category=%s,cost_price=%s,price=%s,sale_price=%s,
                         stock=%s,image_url=%s,description=%s WHERE sku=%s"""),
                    (p.name,p.category,p.cost_price,p.price,p.sale_price,p.stock,p.image_url,p.description,p.sku))
    else:
        cur.execute(q("""INSERT INTO products (sku,name,category,cost_price,price,sale_price,stock,image_url,description)
                         VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)"""),
                    (p.sku,p.name,p.category,p.cost_price,p.price,p.sale_price,p.stock,p.image_url,p.description))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "sku": p.sku}

class BulkProduct(BaseModel):
    sku: str
    name: str
    category: str
    cost_price: int = 0
    price: int = 0
    sale_price: int = 0
    stock: int = 0
    description: str = ""

class BulkUpload(BaseModel):
    products: List[BulkProduct]

@app.post("/api/admin/products/bulk")
def bulk_upload(payload: BulkUpload, x_admin_token: Optional[str] = Header(None)):
    require_admin(x_admin_token)
    conn = get_conn(); cur = conn.cursor()
    added = 0; updated = 0; errors = []
    for i, p in enumerate(payload.products):
        try:
            if not p.sku or not p.name or not p.category:
                errors.append(f"Row {i+1}: missing SKU, name, or category")
                continue
            # sale_price defaults to price if not given
            sale = p.sale_price if p.sale_price > 0 else p.price
            cur.execute(q("SELECT sku FROM products WHERE sku=%s"), (p.sku,))
            if cur.fetchone():
                cur.execute(q("""UPDATE products SET name=%s,category=%s,cost_price=%s,price=%s,sale_price=%s,
                                 stock=%s,description=%s WHERE sku=%s"""),
                            (p.name,p.category,p.cost_price,p.price,sale,p.stock,p.description,p.sku))
                updated += 1
            else:
                cur.execute(q("""INSERT INTO products (sku,name,category,cost_price,price,sale_price,stock,image_url,description)
                                 VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)"""),
                            (p.sku,p.name,p.category,p.cost_price,p.price,sale,p.stock,"",p.description))
                added += 1
        except Exception as e:
            errors.append(f"Row {i+1} ({p.sku}): {str(e)}")
    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "added": added, "updated": updated, "errors": errors}



@app.delete("/api/admin/products/{sku}")
def delete_product(sku: str, x_admin_token: Optional[str] = Header(None)):
    require_admin(x_admin_token)
    conn = get_conn(); cur = conn.cursor()
    cur.execute(q("DELETE FROM products WHERE sku=%s"), (sku,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}


class BulkDelete(BaseModel):
    skus: List[str] = []
    category: Optional[str] = None
    all: bool = False

@app.post("/api/admin/products/bulk-delete")
def bulk_delete_products(payload: BulkDelete, x_admin_token: Optional[str] = Header(None)):
    require_admin(x_admin_token)
    conn = get_conn(); cur = conn.cursor()
    deleted = 0
    if payload.all:
        cur.execute("SELECT COUNT(*) AS c FROM products")
        row = cur.fetchone(); deleted = row["c"] if isinstance(row, dict) else row[0]
        cur.execute("DELETE FROM products")
    elif payload.category:
        cur.execute(q("SELECT COUNT(*) AS c FROM products WHERE category=%s"), (payload.category,))
        row = cur.fetchone(); deleted = row["c"] if isinstance(row, dict) else row[0]
        cur.execute(q("DELETE FROM products WHERE category=%s"), (payload.category,))
    elif payload.skus:
        for sku in payload.skus:
            cur.execute(q("DELETE FROM products WHERE sku=%s"), (sku,))
            deleted += 1
    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "deleted": deleted}


# ------------------------------------------------------- admin: categories
@app.post("/api/admin/categories")
def upsert_category(c: CategoryIn, x_admin_token: Optional[str] = Header(None)):
    require_admin(x_admin_token)
    conn = get_conn(); cur = conn.cursor()
    cur.execute(q("SELECT name FROM categories WHERE name=%s"), (c.name,))
    if cur.fetchone():
        cur.execute(q("UPDATE categories SET sort_order=%s, image_url=%s WHERE name=%s"), (c.sort_order, c.image_url, c.name))
    else:
        cur.execute(q("INSERT INTO categories (name,sort_order,image_url) VALUES (%s,%s,%s)"), (c.name, c.sort_order, c.image_url))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}


@app.delete("/api/admin/categories/{name}")
def delete_category(name: str, x_admin_token: Optional[str] = Header(None)):
    require_admin(x_admin_token)
    conn = get_conn(); cur = conn.cursor()
    cur.execute(q("DELETE FROM categories WHERE name=%s"), (name,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}


# ------------------------------------------------------- admin: settings (offer banner, etc.)
@app.get("/api/admin/settings")
def get_settings(x_admin_token: Optional[str] = Header(None)):
    require_admin(x_admin_token)
    return {
        "banner_text": get_setting("banner_text", ""),
        "banner_enabled": get_setting("banner_enabled", "false") == "true",
    }


@app.post("/api/admin/settings")
def save_settings(s: SettingsIn, x_admin_token: Optional[str] = Header(None)):
    require_admin(x_admin_token)
    conn = get_conn(); cur = conn.cursor()
    for key, val in (("banner_text", s.banner_text), ("banner_enabled", "true" if s.banner_enabled else "false")):
        if USING_PG:
            cur.execute("""INSERT INTO settings (key,value) VALUES (%s,%s)
                           ON CONFLICT (key) DO UPDATE SET value=EXCLUDED.value""", (key, val))
        else:
            cur.execute("INSERT OR REPLACE INTO settings (key,value) VALUES (?,?)", (key, val))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}


# ------------------------------------------------------- footfall / analytics
@app.post("/api/track/visit")
def track_visit():
    conn = get_conn(); cur = conn.cursor()
    cur.execute(q("INSERT INTO analytics_events (created_at,kind,sku) VALUES (%s,%s,%s)"),
                (datetime.now(timezone.utc).isoformat(timespec="seconds"), "visit", None))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}


@app.post("/api/track/view/{sku}")
def track_product_view(sku: str):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(q("INSERT INTO analytics_events (created_at,kind,sku) VALUES (%s,%s,%s)"),
                (datetime.now(timezone.utc).isoformat(timespec="seconds"), "product_view", sku))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}


@app.get("/api/admin/analytics")
def admin_analytics(x_admin_token: Optional[str] = Header(None)):
    require_admin(x_admin_token)
    conn = get_conn(); cur = conn.cursor()

    cur.execute(q("SELECT COUNT(*) AS c FROM analytics_events WHERE kind=%s"), ("visit",))
    row = cur.fetchone()
    total_visits = row["c"] if isinstance(row, dict) else row[0]

    cur.execute(q("SELECT sku, COUNT(*) AS views FROM analytics_events WHERE kind=%s GROUP BY sku ORDER BY views DESC LIMIT 10"), ("product_view",))
    view_rows = [dict(r) for r in cur.fetchall()]

    names = {}
    if view_rows:
        cur.execute("SELECT sku, name FROM products")
        for r in cur.fetchall():
            r = dict(r); names[r["sku"]] = r["name"]

    cur.close(); conn.close()
    top_viewed = [{"sku": r["sku"], "name": names.get(r["sku"], r["sku"]), "views": r["views"]} for r in view_rows]
    return {"total_visits": total_visits, "top_viewed": top_viewed}


# ------------------------------------------------------- admin: orders + reports
@app.get("/api/admin/orders")
def admin_orders(x_admin_token: Optional[str] = Header(None)):
    require_admin(x_admin_token)
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM orders ORDER BY created_at DESC")
    rows = cur.fetchall(); cur.close(); conn.close()
    out = []
    for r in rows:
        d = dict(r); d["items"] = json.loads(d.pop("items_json")); out.append(d)
    return out


@app.post("/api/admin/orders/{order_id}/status")
def order_status(order_id: str, request: Request, x_admin_token: Optional[str] = Header(None)):
    require_admin(x_admin_token)
    status = request.query_params.get("status", "")
    if status not in ("new","confirmed","shipped","delivered","cancelled"):
        raise HTTPException(400, "Invalid status.")
    conn = get_conn(); cur = conn.cursor()
    cur.execute(q("UPDATE orders SET status=%s WHERE id=%s"), (status, order_id))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "status": status}


@app.get("/api/admin/reports")
def reports(x_admin_token: Optional[str] = Header(None)):
    require_admin(x_admin_token)
    conn = get_conn(); cur = conn.cursor()

    cur.execute("SELECT total, cost_total, status FROM orders")
    orders = [dict(r) for r in cur.fetchall()]
    valid = [o for o in orders if o["status"] != "cancelled"]
    revenue = sum(o["total"] for o in valid)
    cogs = sum(o["cost_total"] for o in valid)
    order_count = len(valid)

    af = "FALSE" if USING_PG else "0"
    cur.execute(f"SELECT sku,name,category,cost_price,sale_price,stock FROM products WHERE archived={af}")
    prods = [dict(r) for r in cur.fetchall()]
    stock_value_cost = sum(p["cost_price"]*p["stock"] for p in prods)
    stock_value_retail = sum(p["sale_price"]*p["stock"] for p in prods)
    out_of_stock = [p for p in prods if p["stock"] == 0]
    low_stock = [p for p in prods if 0 < p["stock"] <= 2]

    cur.execute("SELECT items_json, status FROM orders")
    sold = {}
    for r in cur.fetchall():
        r = dict(r)
        if r["status"] == "cancelled":
            continue
        for it in json.loads(r["items_json"]):
            sold[it["sku"]] = sold.get(it["sku"], 0) + it["qty"]
    top_sellers = sorted(
        [{"sku":p["sku"],"name":p["name"],"units":sold.get(p["sku"],0)} for p in prods],
        key=lambda x: x["units"], reverse=True)[:10]

    cur.close(); conn.close()
    return {
        "revenue": revenue, "cogs": cogs, "gross_profit": revenue - cogs,
        "order_count": order_count,
        "avg_order_value": round(revenue/order_count) if order_count else 0,
        "product_count": len(prods),
        "stock_value_cost": stock_value_cost, "stock_value_retail": stock_value_retail,
        "out_of_stock_count": len(out_of_stock),
        "low_stock": low_stock,
        "out_of_stock": [{"sku":p["sku"],"name":p["name"]} for p in out_of_stock],
        "top_sellers": top_sellers,
    }


# ------------------------------------------------------- pages
@app.get("/")
def home(): return FileResponse("index.html")

@app.get("/about")
def page_about(): return FileResponse("page_about.html")

@app.get("/contact")
def page_contact(): return FileResponse("page_contact.html")

@app.get("/refund-policy")
def page_refund(): return FileResponse("page_refund.html")

@app.get("/shipping-policy")
def page_shipping(): return FileResponse("page_shipping.html")

@app.get("/privacy-policy")
def page_privacy(): return FileResponse("page_privacy.html")

@app.get("/terms")
def page_terms(): return FileResponse("page_terms.html")

from fastapi.responses import Response

@app.get("/sitemap.xml")
def sitemap():
    urls = ["", "about", "contact", "refund-policy", "shipping-policy", "privacy-policy", "terms"]
    base = "https://houseofranishe.in/"
    items = "".join(f"<url><loc>{base}{u}</loc><changefreq>weekly</changefreq><priority>{'1.0' if u=='' else '0.7'}</priority></url>" for u in urls)
    xml = f'<?xml version="1.0" encoding="UTF-8"?><urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">{items}</urlset>'
    return Response(content=xml, media_type="application/xml")

@app.get("/robots.txt")
def robots():
    txt = "User-agent: *\nAllow: /\nSitemap: https://houseofranishe.in/sitemap.xml\n"
    return Response(content=txt, media_type="text/plain")


@app.get("/api/admin/reports/export")
def export_reports(x_admin_token: Optional[str] = Header(None)):
    require_admin(x_admin_token)
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    conn = get_conn(); cur = conn.cursor()

    # Gather data
    cur.execute("SELECT * FROM orders ORDER BY created_at DESC")
    orders = [dict(r) for r in cur.fetchall()]
    valid = [o for o in orders if o["status"] != "cancelled"]
    revenue = sum(o["total"] for o in valid)
    cogs = sum(o["cost_total"] for o in valid)

    af = "FALSE" if USING_PG else "0"
    cur.execute(f"SELECT sku,name,category,cost_price,price,sale_price,stock FROM products WHERE archived={af} ORDER BY category,name")
    prods = [dict(r) for r in cur.fetchall()]

    # units sold
    sold = {}
    for o in orders:
        if o["status"]=="cancelled": continue
        for it in json.loads(o["items_json"]):
            sold[it["sku"]] = sold.get(it["sku"],0) + it["qty"]
    cur.close(); conn.close()

    wb = openpyxl.Workbook()
    HEAD = Font(name="Arial", bold=True, color="FFFFFF")
    HFILL = PatternFill("solid", fgColor="3A2A22")
    GFILL = PatternFill("solid", fgColor="B08D57")

    # Sheet 1: Summary / P&L
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = "HOUSE OF RANISHE — BUSINESS REPORT"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = HFILL
    ws.merge_cells("A1:B1")
    ws["A2"] = "Generated"; ws["B2"] = datetime.now(timezone.utc).strftime("%d %b %Y %H:%M UTC")
    rows = [
        ("", ""),
        ("Revenue (Rs)", revenue),
        ("Cost of Goods (Rs)", cogs),
        ("Gross Profit (Rs)", revenue - cogs),
        ("Total Orders", len(valid)),
        ("Avg Order Value (Rs)", round(revenue/len(valid)) if valid else 0),
        ("Products Live", len(prods)),
        ("Stock Value at Cost (Rs)", sum(p["cost_price"]*p["stock"] for p in prods)),
        ("Stock Value at Retail (Rs)", sum(p["sale_price"]*p["stock"] for p in prods)),
        ("Out of Stock Items", sum(1 for p in prods if p["stock"]==0)),
    ]
    r = 4
    for label, val in rows:
        ws.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 22

    # Sheet 2: Products (with units sold)
    ws2 = wb.create_sheet("Products")
    headers = ["SKU","Name","Category","Cost (Rs)","Price (Rs)","Sale (Rs)","Stock","Units Sold"]
    for c,h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=c, value=h); cell.font = HEAD; cell.fill = GFILL
    r = 2
    for p in prods:
        vals = [p["sku"],p["name"],p["category"],p["cost_price"],p["price"],p["sale_price"],p["stock"],sold.get(p["sku"],0)]
        for c,v in enumerate(vals, start=1): ws2.cell(row=r, column=c, value=v)
        r += 1
    widths2 = [12,30,14,12,12,12,10,12]
    for i,w in enumerate(widths2, start=1): ws2.column_dimensions[chr(64+i)].width = w
    ws2.freeze_panes = "A2"

    # Sheet 3: Orders
    ws3 = wb.create_sheet("Orders")
    oh = ["Order ID","Date","Customer","Phone","Items","Total (Rs)","Status"]
    for c,h in enumerate(oh, start=1):
        cell = ws3.cell(row=1, column=c, value=h); cell.font = HEAD; cell.fill = GFILL
    r = 2
    for o in orders:
        items = "; ".join(f"{it['name']} x{it['qty']}" for it in json.loads(o["items_json"]))
        vals = [o["id"], o["created_at"][:16].replace("T"," "), o["customer_name"], o["phone"], items, o["total"], o["status"]]
        for c,v in enumerate(vals, start=1): ws3.cell(row=r, column=c, value=v)
        r += 1
    widths3 = [16,18,20,14,44,12,12]
    for i,w in enumerate(widths3, start=1): ws3.column_dimensions[chr(64+i)].width = w
    ws3.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"ranishe-report-{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.get("/admin")
def admin_page(): return FileResponse("admin.html")
